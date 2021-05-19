import openpyxl
from openpyxl import *

Excel_file = openpyxl.load_workbook("P&L for python.xlsx")  # To read the file
#  print(type(wb))
sheet_names = Excel_file.sheetnames  # To get sheet names
# print(sheets)
print(Excel_file.active.title)  # To get the active sheet
Sheet1 = Excel_file[Excel_file.active.title]  # To store the sheet
print(type(Sheet1))

# Option 1 to read cell data is given below
data1 = Sheet1["C117"].value  # To fetch data from a particular cell
print(f"data1 is {data1}")
data2 = Excel_file["P&L writing"]["c125"].value
print(f"data2 is {data2}")

# Option 2 to read cell data is given below
data3 = Sheet1.cell(128, 3).value  # cell(raw number,column number)
print(f"data3 is {data3}")
data4 = Excel_file["P&L writing"].cell(131, 3).value
print(f"data4 is {data4}")
data5 = openpyxl.load_workbook("P&L for python.xlsx")["P&L writing"].cell(100,3).value
print(f"data5 is {data5}")

# Option3
data6 = Sheet1.cell(row=128, column= 3).value
print(f"data6 is {data6}")

# To extract data from a xlsx and convert it in to a list using while loop
row_start = 72
Data_list2 = []
while row_start < 200:
    column_start = 2
    Data_list = []
    while column_start < 9:
        z = Sheet1.cell(row=row_start, column=column_start).value
        Data_list.append(z)
        column_start += 1
    Data_list2.append(Data_list)
    row_start += 1

test_list = [['Date', 'Company', 'QTY', 'BUY', 'SELL', 'Closing Date', 'DP Applicable?'],
             ['15/12/2020', 'BOROLTD', 10, 196.4, 198.35, '23/04/2021', 'yes'],
             ['21/12/2020', 'ESCORTS', 2, 1301, None, None, None],
             ['22/12/2020', 'ESCORTS', 2, 1230, None, None, None],
             ['24/12/2020', 'ESCORTS', 2, 1245, None, None, None],
             ['28/12/2020', 'ESCORTS', 2, 1245, None, None, None],
             ['31/12/2020', 'BORORENEW', 5, 312.2, None, None, None],
             ['05/01/2021', 'BECTORFOOD', 10, 485.9, None, None, None],
             ['05/01/2021', 'IEX', 10, 221.7, None, None, None],
             ['11/12/2020', 'IRCTC', 2, 1417, None, None, None],
             ['05/01/2021', 'BORORENEW', 10, 284.8, None, None, None],
             ['08/01/2021', 'TOTAL', 25, 80.05, None, None, None],
             ['13/01/2021', 'TOTAL', 5, 68.05, None, None, None],
             ['13/01/2021', 'GLENMARK', 10, 510.1, 513.9, '08/04/2021', 'yes'],
             ['20/01/2021', 'BECTORFOOD', 1, 434.2, None, None, None],
             ['21/12/2020', 'GLENMARK', 4, 502.65, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'GLENMARK', 1, 501.05, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'GLENMARK', 1, 500.1, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'ASIANPAINT', 2, 2742, None, None, None],
             ['21/12/2020', 'ASIANPAINT', 2, 2700, None, None, None],
             ['21/12/2020', 'GRANULES', 5, 340.25, None, None, None],
             ['22/01/2021', 'ASIANPAINT', 2, 2615, None, None, None],
             ['22/01/2021', 'ASIANPAINT', 1, 2592, None, None, None],
             ['25/01/2021', 'ESCORTS', 1, 1245, None, None, None],
             ['25/01/2021', 'ESCORTS', 1, 1225, None, None, None],
             ['27/01/2021', 'ASIANPAINT', 2, 2440, None, None, None],
             ['27/01/2021', 'GLENMARK', 2, 490, 513.9, '08/04/2021', 'Sold w others']]


# To create a new xlsx file from scratch
Excel_file = "P&L for python.xlsx"
Excel_file["Sheet"].title = "P&LLL"
Sheet_1 = Excel_file.active
# to write any list of lists to individual points to excel file
column_start = 1
for i in test_list:
    row_start = 1
    for j in i:
        Sheet_1.cell(row=column_start, column=row_start, value=j)
        row_start = row_start + 1
    column_start = column_start + 1
Excel_file.save("5Python.xlsx")
