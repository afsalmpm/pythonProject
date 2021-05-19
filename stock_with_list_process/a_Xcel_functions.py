import openpyxl
from openpyxl.styles import Alignment
# ********************** Add your excel file path here ********************** #
excel_file_path = ""
#  Example: excel_file_path = "C:\\Users\\dell\\Desktop\\Stock\\Individual Stock files\\"
# ********************** Add your excel file path here ********************** #


def reading_a_cell_from_excel(file_name, sheet_name, row_number, column_number):
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"
    excel_file = openpyxl.load_workbook(excel_file_name)  # To read the file
    sheet_n = excel_file[sheet_name]  # To store the sheet
    cell_value = sheet_n.cell(row=row_number, column=column_number).value
    return cell_value


def over_writing_a_cell_to_excel(file_name, sheet_name, row_number, column_number, value_to_be_written):
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"
    excel_file = openpyxl.load_workbook(excel_file_name)  # To read the file
    sheet_n = excel_file[sheet_name]  # To store the sheet
    sheet_n.cell(row=row_number, column=column_number, value=value_to_be_written)  # To write data
    sheet_n.cell(row=row_number, column=column_number).alignment = Alignment(horizontal='center')
    excel_file.save(excel_file_name)


def reading_list_from_excel(file_name, sheet_name, row_start_for_reading, row_ending_for_reading,
                            column_start_for_reading, column_ending_for_reading):
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"
    excel_file = openpyxl.load_workbook(excel_file_name)  # To read the file
    sheet_n = excel_file[sheet_name]  # To store the sheet
    # To extract data from excel file by nested for loop method
    data_list = []
    for i in range(row_start_for_reading, row_ending_for_reading + 1):
        list_in_data_list = []
        for j in range(column_start_for_reading, column_ending_for_reading + 1):
            cell_value = sheet_n.cell(row=i, column=j).value  # Data extraction form particular cell
            list_in_data_list.append(cell_value)  # To create list of each line
        data_list.append(list_in_data_list)  # To create list of list for further reference and process
    return data_list


def over_write_list_to_excel(file_name, sheet_name, list_to_be_added_for_writing_to_excel, row_start, column_start):
    listt = []
    for i in list_to_be_added_for_writing_to_excel:
        if i is None:
            listt.append("")
        else:
            listt.append(i)
    list_to_be_added_for_writing_to_excel = listt
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"  # Variable to be excel.Workbook object
    excel_file = openpyxl.load_workbook(excel_file_name)
    sheet_n = excel_file[sheet_name]
    for i in list_to_be_added_for_writing_to_excel:
        sheet_n.cell(row=row_start, column=column_start, value=i)  # To write data
        sheet_n.cell(row=row_start, column=column_start).alignment = Alignment(horizontal='center')
        column_start += 1
    excel_file.save(excel_file_name)


def over_write_list_of_list_to_excel(file_name, sheet_name, row_start, column_start,
                                     list_of_list_to_be_added):
    list_of_list = []
    for i in list_of_list_to_be_added:
        listt = []
        for j in i:
            if j is None:
                listt.append("")
            else:
                listt.append(j)
        list_of_list.append(listt)
    list_of_list_to_be_added = list_of_list
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"
    excel_file = openpyxl.load_workbook(excel_file_name)  # To read the file
    sheet_n = excel_file[sheet_name]  # To store the sheet
    for i in list_of_list_to_be_added:
        column_start_in_iteration = column_start
        for j in i:
            sheet_n.cell(row=row_start, column=column_start_in_iteration, value=j)
            sheet_n.cell(row=row_start, column=column_start_in_iteration).alignment = Alignment(
                horizontal='center')  # To make the cell value in the center
            column_start_in_iteration = column_start_in_iteration + 1
        row_start = row_start + 1
    excel_file.save(excel_file_name)


def search_index_of_a_data_from_a_list(data_list, data):
    # print(data_list)
    starting_of_blank_line = data_list.index(data)
    return starting_of_blank_line


def iteration_to_find_a_filled_till(file_name, sheet_name, row_start_for_reading, row_ending_for_reading,
                                    column_start_for_reading, column_ending_for_reading, blank_list_format):
    data_list = reading_list_from_excel(file_name, sheet_name, row_start_for_reading, row_ending_for_reading,
                                        column_start_for_reading, column_ending_for_reading)
    filled_till = search_index_of_a_data_from_a_list(data_list, blank_list_format)
    row_start = filled_till + row_start_for_reading
    return row_start


def to_make_blank_line(file_name, sheet_name, row_line, column_start, column_end):
    excel_file_name = f"{excel_file_path}{file_name}.xlsx"
    excel_file = openpyxl.load_workbook(excel_file_name)  # To read the file
    sheet_n = excel_file[sheet_name]  # To store the sheet
    while column_start < column_end + 1:
        sheet_n.cell(row=row_line, column=column_start, value="")
        column_start += 1
    excel_file.save(excel_file_name)


def to_insert_row_copy_n_move_down(file_name, sheet_name, row_start_for_reading, row_ending_for_reading,
                                   column_start_for_reading, column_ending_for_reading):
    read_list = reading_list_from_excel(file_name, sheet_name, row_start_for_reading, row_ending_for_reading,
                                        column_start_for_reading, column_ending_for_reading)
    to_make_blank_line(file_name, sheet_name, row_start_for_reading, column_start_for_reading,
                       column_ending_for_reading)
    # All "None"s in the read_list should be converted to " "
    list_of_list = []
    for i in read_list:
        listt = []
        for j in i:
            if j is None:
                listt.append("")
            else:
                listt.append(j)
        list_of_list.append(listt)
    read_list = list_of_list
    over_write_list_of_list_to_excel(file_name, sheet_name, row_start_for_reading + 1, column_start_for_reading,
                                     read_list)


#  ****************************************************************************************************** #
#  *****************************************************************************************************  #
#  ***********************************Excel functions finished******************************************  #
#  ****************************************************************************************************** #
#  *****************************************************************************************************  #
# file_name = "C:\\Users\\dell\\Desktop\\P&L Record May 2021"
# sheet_name = "P&L writing"
# x = float(reading_a_cell_from_excel(file_name, sheet_name, 6, 10))
