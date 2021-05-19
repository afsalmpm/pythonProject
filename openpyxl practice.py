import openpyxl
from openpyxl.styles import Alignment

# 1 Accessing the excel file
# 2 Accessing the names of the sheets
# 3 Accessing the cell in a particular sheet (Reading data and writing data)

excel_file = openpyxl.load_workbook("P&L for python.xlsx")  # To read the file
excel_file.sheetnames  # To get all sheet name, You can add print to display
sheet1 = excel_file["P&L writing"]  # To store the sheet


#888888888888888888888888888888888888888888888888888888888888888888888888888888#
# Option 1 to read cell data is given below
data1 = sheet1["C117"].value  # To fetch data from a particular cell
print(f"data1 is {data1}")
data2 = excel_file["P&L writing"]["c125"].value
print(f"data2 is {data2}")
# Option 2 to read cell data is given below
data3 = sheet1.cell(128, 3).value  # cell(raw number,column number)
print(f"data3 is {data3}")
data4 = excel_file["P&L writing"].cell(131, 3).value
print(f"data4 is {data4}")
data5 = openpyxl.load_workbook("C:\\Users\\dell\\Desktop\\P&L for python.xlsx")["P&L writing"].cell(100,3).value
print(f"data5 is {data5}")
# Option3
data6 = sheet1.cell(row=128, column= 3).value
print(f"data6 is {data6}")
#888888888888888888888888888888888888888888888888888888888888888888888888888888#







# To extract data from by nested for loop method

row_start = 73  # Starting of extraction of data from nth row
data_list = []
row_ending = 200  # Ending of extraction of data till nth row

# Data extraction using for loop

for i in range(row_start, row_ending+1):
    column_start = 2  # data extraction column starting number
    list_in_data_list = []
    column_ending = 8  # data extraction column ending number
    for j in range(column_start, column_ending+1):
        cell_value = sheet1.cell(row=i, column=j).value  # Data extraction form particular cell
        list_in_data_list.append(cell_value)  # To create list of each line
    data_list.append(list_in_data_list)  # To create "list of list" for further reference and process


# print(data_list)
# print("\n")
# print("Printing of list of list ended here")

starting_of_blank_line = data_list.index([None, None, None, None, None, None, None])
print(f" The data is currently filled till {starting_of_blank_line}th line")

# Below this need to make a code for adding new trades and to be stored as test_list

test_list = stock_entry()  # Calling the stock entry function from Utils

# test_list = [['15/12/2020', 'BOROLTD', 10, 196.4, 198.35, '23/04/2021', 'yes'],
#              ['21/12/2020', 'ESCORTS', 2, 1301, None, None, None],
#              ['22/12/2020', 'ESCORTS', 2, 1230, None, None, None],
#              ['24/12/2020', 'ESCORTS', 2, 1245, None, None, None],
#              ['28/12/2020', 'ESCORTS', 2, 1245, None, None, None],
#              ['31/12/2020', 'BORORENEW', 5, 312.2, None, None, None],
#              ['05/01/2021', 'BECTORFOOD', 10, 485.9, None, None, None],
#              ['05/01/2021', 'IEX', 10, 221.7, None, None, None],
#              ['11/12/2020', 'IRCTC', 2, 1417, None, None, None],
#              ['05/01/2021', 'BORORENEW', 10, 284.8, None, None, None],
#              ['08/01/2021', 'TOTAL', 25, 80.05, None, None, None],
#              ['13/01/2021', 'TOTAL', 5, 68.05, None, None, None],
#              ['13/01/2021', 'GLENMARK', 10, 510.1, 513.9, '08/04/2021', 'yes'],
#              ['20/01/2021', 'BECTORFOOD', 1, 434.2, None, None, None],
#              ['21/12/2020', 'GLENMARK', 4, 502.65, 513.9, '08/04/2021', 'Sold w others'],
#              ['21/12/2020', 'GLENMARK', 1, 501.05, 513.9, '08/04/2021', 'Sold w others'],
#              ['21/12/2020', 'GLENMARK', 1, 500.1, 513.9, '08/04/2021', 'Sold w others'],
#              ['21/12/2020', 'ASIANPAINT', 2, 2742, None, None, None],
#              ['21/12/2020', 'ASIANPAINT', 2, 2700, None, None, None],
#              ['21/12/2020', 'GRANULES', 5, 340.25, None, None, None],
#              ['22/01/2021', 'ASIANPAINT', 2, 2615, None, None, None],
#              ['22/01/2021', 'ASIANPAINT', 1, 2592, None, None, None],
#              ['25/01/2021', 'ESCORTS', 1, 1245, None, None, None],
#              ['25/01/2021', 'ESCORTS', 1, 1225, None, None, None],
#              ['27/01/2021', 'ASIANPAINT', 2, 2440, None, None, None],
#              ['27/01/2021', 'GLENMARK', 2, 490, 513.9, '08/04/2021', 'Sold w others']]

# To write the manipulated to as new XL
excel_file = openpyxl.load_workbook("P&L for python.xlsx")  # To read the file
#  print(type(wb))
#  sheets = excel_file.sheetnames  # To get sheet names
#  print(sheets)
print(excel_file.active.title)  # To get the active sheet
sheet1 = excel_file[excel_file.active.title]  # To store the sheet
print(type(sheet1))
row_start = 73 + starting_of_blank_line
for i in test_list:
    column_start = 2
    for j in i:
        sheet1.cell(row=row_start, column=column_start, value=j)  # To write data
        sheet1.cell(row=row_start, column=column_start).alignment = Alignment(horizontal='center')  #To make the cell value in the ceter
        column_start = column_start + 1
    row_start = row_start + 1
excel_file.save("dP&L for python.xlsx")