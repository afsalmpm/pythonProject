import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\dell\\Desktop\\P&L for python.xlsx")  # To read the file
#print(type(wb))
#sheets = wb.sheetnames  # To get sheet names
# # print(sheets)
# print(wb.active.title)  # To get the active sheet
sh1 = wb["P&L writing"]  # To store the sheet
# print(type(sh1))
#
# # Option 1 to read cell data is given below
# data1 = sh1["C117"].value  # To fetch data from a particular cell
# print(f"data1 is {data1}")
# data2 = wb["P&L writing"]["c125"].value
# print(f"data2 is {data2}")
#
# # Option 2 to read cell data is given below
# data3 = sh1.cell(128,3).value  # cell(raw number,column number)
# print(f"data3 is {data3}")
# data4 = wb["P&L writing"].cell(131,3).value
# print(f"data4 is {data4}")
# data5 = openpyxl.load_workbook("C:\\Users\\dell\\Desktop\\P&L for python.xlsx")["P&L writing"].cell(100,3).value
# print(f"data5 is {data5}")
#
# # Option3
# data6 = sh1.cell(row=128, column= 3).value
# print(f"data6 is {data6}")

row_start = 72
Data_list2 = []
row_ending = 163

print("Using for loop \n")


for i in range(row_start, row_ending+1):
    column_start = 2
    Data_list = []
    column_ending = 8
    for j in range(column_start, column_ending+1):
        z = sh1.cell(row=i, column=j).value
        Data_list.append(z)
    Data_list2.append(Data_list)
    print(Data_list, "\n")
print("\n")
print("Using while loop \n")
while row_start < row_ending+1:
    column_start = 2
    Data_list = []
    column_ending = 8
    while column_start < column_ending+1:
        z = sh1.cell(row=row_start, column=column_start).value
        Data_list.append(z)
        column_start += 1
    print(Data_list, "\n")
    Data_list2.append(Data_list)
    row_start += 1

sh1.cell(row=164,column=2, value="WTF")
sh1.cell(row=164,column=3, value="WTF")
sh1.cell(row=164,column=4, value="WTF")
sh1.cell(row=164,column=5, value="WTF")

wb.save("C:\\Users\\dell\\Desktop\\P&L for python1.xlsx")







import csv

# data to be written row-wise in csv fil
data = Data_list2

# opening the csv file in 'w+' mode
file = open('April P&L.csv', 'w+', newline='')

# writing the data into the file
with file:
    write = csv.writer(file)
    write.writerows(data)


