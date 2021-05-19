from openpyxl import *


def make_excel_file():
    stock_name = input("Enter the file name:")
    excel_file_name = f"{stock_name}.xlsx"  # Variable to be excel.Workbook object
    excel_file = Workbook()
    count = int(input("Number of sheets required:"))
    for num in range(count):
        excel_file.create_sheet(f"{input(f'Enter the sheet name {num+1}:')}", num)
    #  print(excel_file.sheetnames)
    excel_file.remove(excel_file["Sheet"])  # To remove default sheet named "Sheet"
    print(excel_file.sheetnames)
    excel_file.save(excel_file_name)


make_excel_file()