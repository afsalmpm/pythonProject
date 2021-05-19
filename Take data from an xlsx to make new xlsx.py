import openpyxl
from openpyxl.styles import Alignment

wb = openpyxl.load_workbook("C:\\Users\\dell\\Desktop\\P&L for python.xlsx")  # To read the file
sh1 = wb["P&L writing"]  # To store the sheet

# To extract data from by nested for loop method

row_start = 73  # Starting of extraction of data from nth row
data_list = []
row_ending = 200  # Ending of extraction of data till nth row

print("Using for loop \n")

for i in range(row_start, row_ending+1):
    column_start = 2  # data extraction column starting number
    list_in_data_list = []
    column_ending = 8  # data extraction column ending number
    for j in range(column_start, column_ending+1):
        cell_value = sh1.cell(row=i, column=j).value  # Data extraction form particular cell
        list_in_data_list.append(cell_value)  # To create list of each line
    data_list.append(list_in_data_list)  # To create list of list for further reference and process


print(data_list)
print("\n")
print("Printing of list of list ended here")

starting_of_blank_line = data_list.index([None, None, None, None, None, None, None])
print(starting_of_blank_line)


# Below this need to make a code for adding new trades and to be stored as test_list

test_list = [['15/12/2020', 'BOROLTD', 10, 196.4, 198.35, '23/04/2021', 'yes'],
             ['21/12/2020', 'ESCORTS', 2, 1301, None, None, None],
             ['22/12/2020', 'ESCORTS', 2, 1230, None, None, None],
             ['24/12/2020', 'ESCORTS', 2, 1245, None, None, None],
             ['28/12/2020', 'ESCORTS', 2, 1245, None, None, None],
             ['31/12/2020', 'BORORENEW', 5, 312.2, None, None, None],
             ['05/01/2021', 'BECTORFOOD', 10, 485.9, None, None, None],
             ['05/01/2021', 'IEX', 10, 221.7, None, None, None],
             ['11/12/2020', 'IRCTC', 2, 1417, None, None, None],
             ['05/01/2021', 'BORORENEW', 10, 284.8, None, None, None],
             ['08/01/2021', 'TOTAL', 25, 80.05, None, None, None],
             ['13/01/2021', 'TOTAL', 5, 68.05, None, None, None],
             ['13/01/2021', 'GLENMARK', 10, 510.1, 513.9, '08/04/2021', 'yes'],
             ['20/01/2021', 'BECTORFOOD', 1, 434.2, None, None, None],
             ['21/12/2020', 'GLENMARK', 4, 502.65, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'GLENMARK', 1, 501.05, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'GLENMARK', 1, 500.1, 513.9, '08/04/2021', 'Sold w others'],
             ['21/12/2020', 'ASIANPAINT', 2, 2742, None, None, None],
             ['21/12/2020', 'ASIANPAINT', 2, 2700, None, None, None],
             ['21/12/2020', 'GRANULES', 5, 340.25, None, None, None],
             ['22/01/2021', 'ASIANPAINT', 2, 2615, None, None, None],
             ['22/01/2021', 'ASIANPAINT', 1, 2592, None, None, None],
             ['25/01/2021', 'ESCORTS', 1, 1245, None, None, None],
             ['25/01/2021', 'ESCORTS', 1, 1225, None, None, None],
             ['27/01/2021', 'ASIANPAINT', 2, 2440, None, None, None],
             ['27/01/2021', 'GLENMARK', 2, 490, 513.9, '08/04/2021', 'Sold w others']]

# To save the manipulated to as new XL
wb = openpyxl.load_workbook("C:\\Users\\dell\\Desktop\\P&L for python.xlsx")  # To read the file
#  print(type(wb))
sheets = wb.sheetnames  # To get sheet names
# print(sheets)
print(wb.active.title)  # To get the active sheet
sh1 = wb[wb.active.title]  # To store the sheet
print(type(sh1))
row_start = 73 + starting_of_blank_line
for i in test_list:
    column_start = 2
    for j in i:
        sh1.cell(row=row_start, column=column_start, value=j)
        sh1.cell(row=row_start, column=column_start).alignment = Alignment(horizontal='center')  # To make the cell value in the ceter
        column_start = column_start + 1
    row_start = row_start + 1
wb.save("C:\\Users\\dell\\Desktop\\savedP&L for python.xlsx")